#!/usr/bin/env python3
"""Export MINI GT workbook rows and in-cell images to static site assets.

Usage:
    python scripts/extract_catalog.py "C:\\path\\to\\Mini GT.xlsx"
"""

from __future__ import annotations

import json
import re
import shutil
import sys
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
ASSETS = ROOT / "assets"
IMAGE_DIR = ASSETS / "images"
DATA_FILE = ASSETS / "data" / "models.json"


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def display_brand(value: Any) -> str:
    """Apply display names that are consistent across every future export."""
    brand = clean_text(value)
    return "FERRARI" if brand.upper() == "FERRARI (BBR)" else brand


def slugify(value: str) -> str:
    normalized = value.lower().replace("★", "star")
    normalized = re.sub(r"[^a-z0-9]+", "-", normalized)
    return normalized.strip("-") or "prototype"


def normalize_date(value: Any, raw_value: str) -> str:
    """Keep true Excel dates, while preserving standalone year values like 2026."""
    raw_value = clean_text(raw_value)
    if re.fullmatch(r"20\d{2}", raw_value):
        return raw_value
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return raw_value or clean_text(value)


def collection_names(model: str) -> list[str]:
    lowered = model.lower()
    names: list[str] = []
    if "fast & furious" in lowered:
        names.append("Fast & Furious Collection")
    if "bond 007" in lowered or "bond 007 series" in lowered or "bond collection" in lowered:
        names.append("Bond 007 Collection")
    if "korean collection" in lowered:
        names.append("Korean Collection")

    for match in re.findall(r"\(([^)]*?collection)\)", model, flags=re.IGNORECASE):
        name = clean_text(match)
        if name.lower() in {"bond 007 collection", "bond 007 series", "most likely bond collection"}:
            name = "Bond 007 Collection"
        elif name.lower() == "fast & furious collection":
            name = "Fast & Furious Collection"
        elif name.lower() == "korean collection":
            name = "Korean Collection"
        if name and name not in names:
            names.append(name)
    return names


def child_by_name(element: ET.Element, name: str) -> ET.Element | None:
    return next((child for child in element if local_name(child.tag) == name), None)


def xml_cell_metadata(archive: ZipFile) -> tuple[list[int], list[str]]:
    """Return value-metadata → rich-value index → embedded image package path."""
    metadata = ET.fromstring(archive.read("xl/metadata.xml"))
    future_metadata = next(
        element for element in metadata.iter() if local_name(element.tag) == "futureMetadata"
    )
    future_to_rich_value: list[int] = []
    for block in future_metadata:
        rich_value = next(
            (node for node in block.iter() if local_name(node.tag) == "rvb"), None
        )
        future_to_rich_value.append(int(rich_value.attrib["i"]) if rich_value is not None else -1)

    value_metadata = next(
        element for element in metadata.iter() if local_name(element.tag) == "valueMetadata"
    )
    metadata_to_rich_value: list[int] = []
    for block in value_metadata:
        rich_cell = next((node for node in block.iter() if local_name(node.tag) == "rc"), None)
        future_index = int(rich_cell.attrib["v"]) if rich_cell is not None else -1
        metadata_to_rich_value.append(
            future_to_rich_value[future_index]
            if 0 <= future_index < len(future_to_rich_value)
            else -1
        )

    rich_value_relationships = ET.fromstring(archive.read("xl/richData/richValueRel.xml"))
    relationship_ids = [
        element.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
        for element in rich_value_relationships
    ]
    package_relationships = ET.fromstring(
        archive.read("xl/richData/_rels/richValueRel.xml.rels")
    )
    relationship_targets = {
        element.attrib["Id"]: element.attrib["Target"].replace("../", "xl/")
        for element in package_relationships
    }

    rich_values = ET.fromstring(archive.read("xl/richData/rdrichvalue.xml"))
    image_paths: list[str] = []
    for rich_value in rich_values:
        first_value = next((node for node in rich_value if local_name(node.tag) == "v"), None)
        relationship_index = int(first_value.text) if first_value is not None else -1
        if 0 <= relationship_index < len(relationship_ids):
            image_paths.append(relationship_targets[relationship_ids[relationship_index]])
        else:
            image_paths.append("")
    return metadata_to_rich_value, image_paths


def raw_cells(archive: ZipFile, sheet_index: int) -> dict[str, tuple[str, int | None]]:
    """Read raw spreadsheet values plus the `vm` reference used by Excel cell images."""
    root = ET.fromstring(archive.read(f"xl/worksheets/sheet{sheet_index}.xml"))
    values: dict[str, tuple[str, int | None]] = {}
    for cell in (node for node in root.iter() if local_name(node.tag) == "c"):
        coordinate = cell.attrib.get("r")
        if not coordinate:
            continue
        value_node = child_by_name(cell, "v")
        raw_value = value_node.text if value_node is not None and value_node.text else ""
        metadata_index = int(cell.attrib["vm"]) if "vm" in cell.attrib else None
        values[coordinate] = (raw_value, metadata_index)
    return values


def image_role(header: str, column_index: int) -> str:
    normalized = header.lower()
    if "real" in normalized or "bollide" in normalized:
        return "realCar"
    return "model" if column_index == 4 else "additional"


def is_cancelled_row(sheet: Any, row_index: int) -> bool:
    """The source workbook marks cancelled prototypes with its yellow theme fill."""
    for column_index in range(1, sheet.max_column + 1):
        fill = sheet.cell(row_index, column_index).fill
        if (
            fill.fill_type == "solid"
            and fill.fgColor.type == "theme"
            and fill.fgColor.theme == 7
        ):
            return True
    return False


def copy_image(archive: ZipFile, package_path: str, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with archive.open(package_path) as source, destination.open("wb") as target:
        shutil.copyfileobj(source, target)


def main(source_arg: str) -> None:
    source = Path(source_arg).expanduser().resolve()
    if not source.is_file():
        raise SystemExit(f"Workbook not found: {source}")

    workbook = load_workbook(source, data_only=True)
    if IMAGE_DIR.exists():
        shutil.rmtree(IMAGE_DIR)
    IMAGE_DIR.mkdir(parents=True)
    DATA_FILE.parent.mkdir(parents=True, exist_ok=True)

    models: list[dict[str, Any]] = []
    missing_images = 0
    unknown_counter = 0

    with ZipFile(source) as archive:
        metadata_to_rich_value, image_paths = xml_cell_metadata(archive)

        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            raw_sheet_cells = raw_cells(archive, sheet_index)
            headers = {
                column: clean_text(sheet.cell(1, column).value)
                for column in range(1, sheet.max_column + 1)
            }
            brand = display_brand(sheet.title)

            for row_index in range(2, sheet.max_row + 1):
                raw_model = clean_text(sheet.cell(row_index, 1).value)
                event = clean_text(sheet.cell(row_index, 2).value)
                date_cell = sheet.cell(row_index, 3)
                raw_date = raw_sheet_cells.get(date_cell.coordinate, ("", None))[0]
                shown_date = normalize_date(date_cell.value, raw_date)

                image_refs: list[dict[str, str]] = []
                for column_index in range(4, sheet.max_column + 1):
                    coordinate = sheet.cell(row_index, column_index).coordinate
                    _, metadata_index = raw_sheet_cells.get(coordinate, ("", None))
                    if metadata_index is None:
                        continue

                    # `vm` is a one-based reference to the <valueMetadata> block.
                    metadata_position = metadata_index - 1
                    if not 0 <= metadata_position < len(metadata_to_rich_value):
                        missing_images += 1
                        continue
                    rich_value_index = metadata_to_rich_value[metadata_position]
                    package_path = (
                        image_paths[rich_value_index]
                        if 0 <= rich_value_index < len(image_paths)
                        else ""
                    )
                    if not package_path:
                        missing_images += 1
                        continue

                    role = image_role(headers[column_index], column_index)
                    image_refs.append({"role": role, "package_path": package_path})

                if not raw_model and not image_refs:
                    continue
                if not raw_model:
                    unknown_counter += 1
                    raw_model = f"Unidentified prototype #{unknown_counter}"

                model_id = f"{slugify(brand)}-{row_index:02d}"
                photos: list[dict[str, str]] = []
                role_counters: dict[str, int] = {}
                for image_index, image_ref in enumerate(image_refs, start=1):
                    role = image_ref["role"]
                    role_counters[role] = role_counters.get(role, 0) + 1
                    suffix = {
                        "model": "model",
                        "additional": "detail",
                        "realCar": "real-car",
                    }[role]
                    if role_counters[role] > 1:
                        suffix = f"{suffix}-{role_counters[role]}"
                    filename = f"{model_id}-{suffix}.png"
                    copy_image(archive, image_ref["package_path"], IMAGE_DIR / filename)
                    label = {
                        "model": "MINI GT model",
                        "additional": "Additional photo",
                        "realCar": "Real car",
                    }[role]
                    photos.append(
                        {
                            "src": f"assets/images/{filename}",
                            "role": role,
                            "label": label,
                        }
                    )

                models.append(
                    {
                        "id": model_id,
                        "name": raw_model,
                        "brand": brand,
                        "event": event,
                        "date": shown_date,
                        "collections": collection_names(raw_model),
                        "cancelled": is_cancelled_row(sheet, row_index),
                        "photos": photos,
                    }
                )

    DATA_FILE.write_text(
        json.dumps(models, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    print(
        f"Exported {len(models)} models, {sum(len(model['photos']) for model in models)} images, "
        f"and {sum(model['cancelled'] for model in models)} cancelled prototypes to {ROOT}"
    )
    if missing_images:
        print(f"Warning: {missing_images} image reference(s) could not be resolved.")


if __name__ == "__main__":
    main(sys.argv[1])
